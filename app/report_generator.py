"""
app/report_generator.py
-----------------------
Generate laporan P&L, Trial Balance summary, dan export ke Excel.
"""

import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Urutan bulan untuk sorting ──────────────────────────────────────────────
MONTH_ORDER = {
    "January 2024": 1, "February 2024": 2, "March 2024": 3,
    "April 2024": 4, "May 2024": 5, "June 2024": 6,
    "July 2024": 7, "August 2024": 8, "September 2024": 9,
    "October 2024": 10, "November 2024": 11, "December 2024": 12,
    # Tambah tahun lain jika perlu
}


def build_pl_summary(df: pd.DataFrame) -> pd.DataFrame:
    """
    Buat summary P&L per bulan dari data gabungan.
    
    Returns:
        DataFrame dengan kolom:
        month | total_revenue | total_cogs | gross_profit | total_opex |
        ebitda | total_other_exp | net_profit | gross_margin | net_margin
    """
    results = []

    for month in df["month"].unique():
        dm = df[df["month"] == month]

        revenue    = dm[dm["category"] == "Revenue"]["amount"].sum()
        cogs       = dm[dm["category"] == "Cost of Goods Sold"]["amount"].sum()
        opex       = dm[dm["category"] == "Operating Expense"]["amount"].sum()
        other_exp  = dm[dm["category"] == "Other Expense"]["amount"].sum()

        gross_profit = revenue - cogs
        ebitda       = gross_profit - opex
        net_profit   = ebitda - other_exp

        gross_margin = (gross_profit / revenue * 100) if revenue > 0 else 0
        net_margin   = (net_profit / revenue * 100) if revenue > 0 else 0

        # KPI Hospitality
        labor_cost = dm[dm["subcategory"] == "Labor Cost"]["amount"].sum()
        food_cost  = dm[dm["subcategory"] == "Food Cost"]["amount"].sum()
        bev_cost   = dm[dm["subcategory"] == "Beverage Cost"]["amount"].sum()
        fnb_rev    = dm[dm["subcategory"] == "Food & Beverage Revenue"]["amount"].sum()

        food_cost_pct  = (food_cost / fnb_rev * 100) if fnb_rev > 0 else 0
        labor_cost_pct = (labor_cost / revenue * 100) if revenue > 0 else 0

        results.append({
            "month": month,
            "total_revenue": revenue,
            "total_cogs": cogs,
            "gross_profit": gross_profit,
            "total_opex": opex,
            "ebitda": ebitda,
            "total_other_expense": other_exp,
            "net_profit": net_profit,
            "gross_margin_%": round(gross_margin, 2),
            "net_margin_%": round(net_margin, 2),
            "food_cost_%": round(food_cost_pct, 2),
            "labor_cost_%": round(labor_cost_pct, 2),
        })

    summary = pd.DataFrame(results)

    # Sort berdasarkan urutan bulan
    summary["_sort"] = summary["month"].map(MONTH_ORDER).fillna(99)
    summary = summary.sort_values("_sort").drop(columns=["_sort"])

    return summary.reset_index(drop=True)


def build_expense_breakdown(df: pd.DataFrame) -> pd.DataFrame:
    """Breakdown biaya per subcategory per bulan."""
    expense_cats = ["Cost of Goods Sold", "Operating Expense", "Other Expense"]
    df_exp = df[df["category"].isin(expense_cats)]

    breakdown = (
        df_exp.groupby(["month", "category", "subcategory"])["amount"]
        .sum()
        .reset_index()
        .rename(columns={"amount": "total_amount"})
    )
    breakdown["_sort"] = breakdown["month"].map(MONTH_ORDER).fillna(99)
    breakdown = breakdown.sort_values(["_sort", "category", "subcategory"]).drop(columns=["_sort"])

    return breakdown.reset_index(drop=True)


def build_general_ledger(df: pd.DataFrame) -> pd.DataFrame:
    """General Ledger view: semua transaksi bersih per akun per bulan."""
    gl = (
        df.groupby(["month", "account_code", "account_name", "category", "subcategory"])
        .agg(debit=("debit", "sum"), credit=("credit", "sum"), amount=("amount", "sum"))
        .reset_index()
    )
    gl["_sort"] = gl["month"].map(MONTH_ORDER).fillna(99)
    gl = gl.sort_values(["_sort", "category", "account_code"]).drop(columns=["_sort"])
    return gl.reset_index(drop=True)


# ── Excel Export ─────────────────────────────────────────────────────────────

def _style_header_row(ws, row_num: int, fill_color: str = "1F3864"):
    """Apply header styling ke satu row di worksheet."""
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[row_num]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def _auto_width(ws):
    """Auto-fit lebar kolom berdasarkan konten."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value)) if cell.value else 0
                max_len = max(max_len, val_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _format_currency_col(ws, col_idx: int, start_row: int, end_row: int):
    """Format kolom sebagai currency Rupiah."""
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal="right")


def export_to_excel(
    pl_summary: pd.DataFrame,
    expense_breakdown: pd.DataFrame,
    general_ledger: pd.DataFrame,
    output_path: str
):
    """
    Export semua laporan ke satu file Excel multi-sheet.
    
    Sheets:
        1. P&L Summary
        2. Expense Breakdown
        3. General Ledger
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Sheet 1: P&L Summary
        pl_summary.to_excel(writer, sheet_name="P&L Summary", index=False)

        # Sheet 2: Expense Breakdown
        expense_breakdown.to_excel(writer, sheet_name="Expense Breakdown", index=False)

        # Sheet 3: General Ledger
        general_ledger.to_excel(writer, sheet_name="General Ledger", index=False)

    # Post-processing: styling dengan openpyxl
    wb = load_workbook(output_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        _style_header_row(ws, 1)
        _auto_width(ws)

        # Freeze panes di baris 2
        ws.freeze_panes = "A2"

        # Format currency: deteksi kolom amount/revenue/cost/profit
        currency_keywords = ["revenue", "cost", "profit", "opex", "expense", "amount", "debit", "credit", "ebitda"]
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value and any(kw in str(cell.value).lower() for kw in currency_keywords):
                _format_currency_col(ws, col_idx, 2, ws.max_row)

        # Zebra stripe rows
        light_fill = PatternFill("solid", fgColor="EEF2F7")
        for row_num in range(2, ws.max_row + 1):
            if row_num % 2 == 0:
                for cell in ws[row_num]:
                    if cell.fill.patternType == "none" or cell.fill.fgColor.rgb == "00000000":
                        cell.fill = light_fill

    wb.save(output_path)
    print(f"  ✓ Excel saved: {output_path}")
